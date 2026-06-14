"""
dca_report.py — DCA 策略 Excel 輸出（三個分頁）。

Sheet1  每日帳戶明細
Sheet2  買入交易紀錄
Sheet3  績效摘要
"""

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dca_config import DCA_MONTHLY_AMOUNT, RISK_FREE_RATE_ANNUAL


# ── 樣式工具 ──────────────────────────────────────────────────────────────────

def _bdr() -> Border:
    s = Side(style="thin", color="B0C4DE")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row: int, col: int, val, width: float | None = None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", start_color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _bdr()
    if width is not None:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _cell(ws, row: int, col: int, val,
          fill: str | None = None,
          bold: bool = False,
          color: str = "000000",
          align: str = "center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=10, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _bdr()
    if fill:
        c.fill = PatternFill("solid", start_color=fill)
    return c


# ── 主函式 ────────────────────────────────────────────────────────────────────

def save_dca_excel(
    output_path:  str,
    display_name: str,
    daily_df,
    trades:       list,
    perf:         dict,
) -> None:
    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet1：每日帳戶明細
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "每日帳戶明細"

    hdrs1   = ["Date", "Close", "累計投入", "持有股數", "帳戶市值", "未實現損益", "報酬率%", "當日新增投入"]
    widths1 = [14,      12,      16,         12,         16,         16,           12,        14]
    for ci, (h, w) in enumerate(zip(hdrs1, widths1), 1):
        _hdr(ws1, 1, ci, h, w)
    ws1.row_dimensions[1].height = 20
    ws1.freeze_panes = "A2"

    buy_dates = {t["buy_date"].strftime("%Y-%m-%d") for t in trades}

    for ri, row in daily_df.iterrows():
        er      = ri + 2
        pnl     = row["未實現損益"]
        ret     = row["報酬率%"]
        inflow  = row["當日新增投入"]
        is_buy  = row["Date"] in buy_dates
        rfill   = "FFD700" if is_buy else ("EBF3FB" if ri % 2 == 0 else None)
        pc = "1F7A3A" if pnl > 0 else ("CC0000" if pnl < 0 else "000000")
        rc = "1F7A3A" if ret > 0 else ("CC0000" if ret < 0 else "000000")

        _cell(ws1, er, 1, row["Date"],      fill=rfill, bold=is_buy, align="center")
        _cell(ws1, er, 2, row["Close"],     fill=rfill, bold=is_buy, align="right")
        _cell(ws1, er, 3, row["累計投入"],  fill=rfill, bold=is_buy, align="right")
        _cell(ws1, er, 4, row["持有股數"],  fill=rfill, bold=is_buy, align="right")
        _cell(ws1, er, 5, row["帳戶市值"],  fill=rfill, bold=is_buy, align="right")
        _cell(ws1, er, 6, pnl,              fill=rfill, bold=is_buy, color=pc, align="right")
        _cell(ws1, er, 7, f"{ret:.2f}%",   fill=rfill, bold=is_buy, color=rc, align="right")
        _cell(ws1, er, 8, inflow if inflow > 0 else "",
                                            fill=rfill, bold=is_buy, align="right")

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet2：買入交易紀錄
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("買入交易紀錄")

    hdrs2   = ["#", "買入日期", "買入價", "股數", "投入成本",
               "期末清算價", "期末市值", "損益", "報酬率%", "持有天數"]
    widths2 = [5,    14,        12,       8,      14,
               14,    14,        14,       12,      12]
    for ci, (h, w) in enumerate(zip(hdrs2, widths2), 1):
        _hdr(ws2, 1, ci, h, w)
    ws2.row_dimensions[1].height = 20
    ws2.freeze_panes = "A2"

    for ri, t in enumerate(trades, 1):
        er   = ri + 1
        pnl  = t["pnl"]
        fill = "C8E6C9" if pnl > 0 else ("FFCDD2" if pnl < 0 else None)
        pc   = "1F7A3A" if pnl > 0 else ("CC0000" if pnl < 0 else "000000")
        vals = [
            t["trade_no"],
            t["buy_date"].strftime("%Y-%m-%d"),
            t["buy_price"],
            t["shares"],
            t["cost"],
            t["sell_price"],
            t["revenue"],
            t["pnl"],
            f"{t['ret_pct']:.2f}%",
            t["hold_days"],
        ]
        alns = ["center","center","right","center","right",
                "right","right","right","right","center"]
        for ci, (v, a) in enumerate(zip(vals, alns), 1):
            color = pc if ci in (8, 9) else "000000"
            _cell(ws2, er, ci, v, fill=fill, color=color, align=a)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet3：績效摘要
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("績效摘要")
    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 20

    cagr_str    = f"{perf['cagr'] * 100:.2f}%"  if perf.get("cagr")    is not None else "N/A"
    irr_str     = f"{perf['irr_val']}%"          if perf.get("irr_val") is not None else "N/A"
    mdd_str     = f"{perf['max_drawdown'] * 100:.2f}%"
    cr_str      = f"{perf['cumulative_ret'] * 100:.2f}%"
    sortino_val = perf.get("sortino_ratio")

    summary = [
        ("每月定期投入金額",                   DCA_MONTHLY_AMOUNT),
        ("策略期間",                           perf["date_range"]),
        ("", ""),
        ("─── 投入與持倉 ───",                ""),
        ("總投入月數（買入次數）",              perf["total_months"]),
        ("累計已投入金額",                      perf["total_invested"]),
        ("未動用零頭現金",                      perf["cash_residual"]),
        ("累計持有股數",                        perf["total_shares"]),
        ("平均買入成本（元/股）",               perf["avg_cost"]),
        ("期末收盤價",                          perf["last_price"]),
        ("", ""),
        ("─── 損益 ───",                       ""),
        ("期末帳戶市值（持股 + 現金）",         perf["final_value"]),
        ("未實現總損益",                        perf["total_pnl"]),
        ("累積報酬率（非年化）",                cr_str),
        ("", ""),
        ("─── 年化報酬率 ───",                 ""),
        ("CAGR（複合年化成長率）",              cagr_str),
        ("年化 IRR",                            irr_str),
        ("", ""),
        ("─── 風險指標 ───",                   ""),
        ("最大回撤 MDD",                        mdd_str),
        (f"Sortino Ratio（無風險利率 {RISK_FREE_RATE_ANNUAL*100:.1f}%，已扣除現金流）",
            sortino_val if sortino_val is not None else "N/A"),
    ]

    title_c = ws3.cell(row=1, column=1, value=f"{display_name} DCA 策略 績效摘要")
    title_c.font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
    title_c.alignment = Alignment(horizontal="center")
    ws3.merge_cells("A1:B1")
    ws3.row_dimensions[1].height = 28

    section_fill = "D6E4F0"
    for ri, (label, val) in enumerate(summary, 2):
        if not label:
            continue
        is_section = label.startswith("───")

        lc = ws3.cell(row=ri, column=1, value=label)
        lc.font      = Font(name="Arial", bold=True, size=10,
                            color="1F4E79" if is_section else "000000")
        lc.fill      = PatternFill("solid",
                                   start_color=section_fill if is_section else "EBF3FB")
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border    = _bdr()

        if is_section:
            vc = ws3.cell(row=ri, column=2, value="")
            vc.fill   = PatternFill("solid", start_color=section_fill)
            vc.border = _bdr()
            ws3.row_dimensions[ri].height = 16
            continue

        if isinstance(val, str) and val not in ("N/A", ""):
            try:
                num = float(val.replace("%", ""))
                tc  = "1F7A3A" if num > 0 else ("CC0000" if num < 0 else "000000")
            except ValueError:
                tc = "000000"
        elif isinstance(val, (int, float)):
            tc = "1F7A3A" if val > 0 else ("CC0000" if val < 0 else "000000")
        else:
            tc = "000000"

        vc = ws3.cell(row=ri, column=2, value=val)
        vc.font      = Font(name="Arial", size=10, color=tc)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border    = _bdr()
        ws3.row_dimensions[ri].height = 18

    wb.save(output_path)
    print(f"  ✅ DCA Excel → {output_path}")
