"""
compare_strategies.py — 趨勢線 vs MACD vs DCA 策略 CAGR / IRR 對比
讀取 test2_output、macd_output、test_3 output 三個資料夾的所有 Excel，
從「績效摘要」Sheet 抓取 CAGR 與 IRR（純已實現平倉版），輸出對比表格。

Sheet1：三策略總覽（CAGR + IRR 兩組指標）
Sheet2：趨勢線 vs MACD（CAGR）
Sheet3：趨勢線 vs DCA（CAGR）
Sheet4：MACD vs DCA（CAGR）
Sheet5：趨勢線 vs MACD（IRR）
Sheet6：趨勢線 vs DCA（IRR）
Sheet7：MACD vs DCA（IRR）

執行方式：
  python compare_strategies.py
"""

import os
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 路徑設定 ──────────────────────────────────────────────────────────────────
DESKTOP     = os.path.join(os.path.expanduser("~"), "Desktop")
TL_DIR      = os.path.join(DESKTOP, "test2_output")
MACD_DIR    = os.path.join(DESKTOP, "macd_output")
DCA_DIR     = os.path.join(DESKTOP, "test_3 output")
OUTPUT_PATH = os.path.join(DESKTOP, "strategy_compare.xlsx")

# ── 樣式常數 ──────────────────────────────────────────────────────────────────
RED_FILL    = PatternFill("solid", start_color="FF4444")
GREY_FILL   = PatternFill("solid", start_color="D0D0D0")
TIE_FILL    = PatternFill("solid", start_color="FFE066")
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
ALT_FILL    = PatternFill("solid", start_color="EBF3FB")
WIN_FONT    = Font(name="Arial", bold=True,  color="FFFFFF", size=11)
LOSE_FONT   = Font(name="Arial", bold=False, color="555555", size=11)
HDR_FONT    = Font(name="Arial", bold=True,  color="FFFFFF", size=11)
NORMAL_FONT = Font(name="Arial", size=11)


def _bdr():
    s = Side(style="thin", color="B0C4DE")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, val, fill=None, font=None, align="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _bdr()
    c.fill      = fill  if fill  else PatternFill()
    c.font      = font  if font  else NORMAL_FONT
    return c


def _sheet_header(ws, col_widths, headers, freeze="A2"):
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = freeze
    for ci, h in enumerate(headers, 1):
        _cell(ws, 1, ci, h, fill=HEADER_FILL, font=HDR_FONT)


def _summary_block(ws, start_row, merge_cols, totals: list[tuple[str, int]]):
    """
    在 start_row 輸出戰果統計小區塊。
    totals = [(label, count), ...]，自動計算勝率，最多勝場標紅。
    merge_cols：標題列合併的欄數（字串，如 "A{r}:D{r}"）
    """
    r = start_row
    ws.row_dimensions[r].height = 22
    title_c = ws.cell(row=r, column=1, value="總戰果")
    title_c.font      = Font(name="Arial", bold=True, size=12, color="1F4E79")
    title_c.alignment = Alignment(horizontal="center", vertical="center")
    title_c.border    = _bdr()
    ws.merge_cells(merge_cols.format(r=r))

    sr = r + 1
    for ci, h in enumerate(["策略", "勝場數", "勝率"], 1):
        c = ws.cell(row=sr, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _bdr()

    total    = sum(cnt for _, cnt in totals)
    non_tie  = [cnt for lbl, cnt in totals if lbl != "平手"]
    max_wins = max(non_tie) if non_tie else 0

    for di, (label, count) in enumerate(totals, sr + 1):
        pct       = f"{count / total * 100:.1f}%" if total else "—"
        is_winner = (label != "平手" and count == max_wins and count > 0)
        sfill = RED_FILL if is_winner else None
        sfont = WIN_FONT if is_winner else NORMAL_FONT
        _cell(ws, di, 1, label, fill=sfill, font=sfont)
        _cell(ws, di, 2, count, fill=sfill, font=sfont)
        _cell(ws, di, 3, pct,   fill=sfill, font=sfont)
        ws.row_dimensions[di].height = 20


# ── 從 Excel 績效摘要 Sheet 抓 CAGR / IRR ───────────────────────────────────

def read_cagr(filepath: str) -> float | None:
    try:
        wb    = load_workbook(filepath, data_only=True)
        sheet = None
        for name in wb.sheetnames:
            if "績效" in name or "摘要" in name:
                sheet = wb[name]
                break
        if sheet is None:
            return None
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "CAGR" in cell.value and "複合" in cell.value:
                    val = sheet.cell(row=cell.row, column=cell.column + 1).value
                    if val is None:
                        return None
                    if isinstance(val, (int, float)):
                        return float(val) * 100
                    if isinstance(val, str):
                        try:
                            return float(val.replace("%", "").strip())
                        except ValueError:
                            return None
        return None
    except Exception as e:
        print(f"  ⚠️  讀取失敗：{filepath}  ({e})")
        return None


def read_irr(filepath: str) -> float | None:
    """
    從「績效摘要」Sheet 抓取「純已實現平倉」版 IRR。
    篩選邏輯：找包含 "IRR" 且不含 "含未平倉" 的標籤列。
    DCA 只有一個 IRR 欄位，同樣能命中。
    """
    try:
        wb    = load_workbook(filepath, data_only=True)
        sheet = None
        for name in wb.sheetnames:
            if "績效" in name or "摘要" in name:
                sheet = wb[name]
                break
        if sheet is None:
            return None
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    v = cell.value
                    if "IRR" in v and "含未平倉" not in v:
                        val = sheet.cell(row=cell.row, column=cell.column + 1).value
                        if val is None:
                            return None
                        if isinstance(val, (int, float)):
                            return float(val) * 100
                        if isinstance(val, str):
                            try:
                                return float(val.replace("%", "").strip())
                            except ValueError:
                                return None
        return None
    except Exception as e:
        print(f"  ⚠️  IRR 讀取失敗：{filepath}  ({e})")
        return None


# ── 掃描資料夾 ────────────────────────────────────────────────────────────────

def scan_folder(folder: str, suffix: str) -> tuple[dict, dict]:
    """
    回傳 (cagr_data, irr_data)，各為 {stock: float | None} 字典。
    """
    cagr_result = {}
    irr_result  = {}
    if not os.path.isdir(folder):
        print(f"  ⚠️  資料夾不存在：{folder}")
        return cagr_result, irr_result
    for path in sorted(glob.glob(os.path.join(folder, "*.xlsx"))):
        basename = os.path.splitext(os.path.basename(path))[0]
        stock    = (basename
                    .replace(suffix, "")
                    .replace("_交易明細", "")
                    .replace("_MACD", "")
                    .replace("_DCA明細", ""))
        cagr = read_cagr(path)
        irr  = read_irr(path)
        cagr_result[stock] = cagr
        irr_result[stock]  = irr
        cagr_str = f"{cagr:.2f}%" if cagr is not None else "N/A"
        irr_str  = f"{irr:.2f}%"  if irr  is not None else "N/A"
        print(f"    {stock:15s}  CAGR = {cagr_str:10s}  IRR = {irr_str}")
    return cagr_result, irr_result


# ── Sheet 建構：三策略總覽 ────────────────────────────────────────────────────

def build_overview_sheet(wb, all_stocks,
                         tl_data, macd_data, dca_data,
                         tl_irr,  macd_irr,  dca_irr):
    ws = wb.active
    ws.title = "三策略總覽"

    _sheet_header(ws,
                  col_widths=[14, 20, 20, 20, 16, 20, 20, 20, 16],
                  headers=["股票代碼",
                           "趨勢線 CAGR", "MACD CAGR", "DCA CAGR", "CAGR 勝出",
                           "趨勢線 IRR",  "MACD IRR",  "DCA IRR",  "IRR 勝出"])

    def _pick_winner(vals: dict) -> tuple[str, str]:
        """
        vals = {"TL": float|None, "MACD": float|None, "DCA": float|None}
        回傳 (winner_key, winner_label)，資料不足回傳 ("N/A", "N/A")
        """
        valid = {k: v for k, v in vals.items() if v is not None}
        if len(valid) < 2:
            return "N/A", "N/A"
        mx      = max(valid.values())
        winners = [k for k, v in valid.items() if v == mx]
        if len(winners) > 1:
            return "平手", "平手"
        return winners[0], {"TL": "趨勢線", "MACD": "MACD", "DCA": "DCA"}[winners[0]]

    tl_w = macd_w = dca_w = ties = 0
    tl_wi = macd_wi = dca_wi = ties_i = 0
    rows = []

    for stock in all_stocks:
        tl_c, macd_c, dca_c = tl_data.get(stock), macd_data.get(stock), dca_data.get(stock)
        tl_i, macd_i, dca_i = tl_irr.get(stock),  macd_irr.get(stock),  dca_irr.get(stock)

        wk, winner    = _pick_winner({"TL": tl_c,  "MACD": macd_c, "DCA": dca_c})
        wki, winner_i = _pick_winner({"TL": tl_i,  "MACD": macd_i, "DCA": dca_i})

        if winner == "平手":       ties   += 1
        elif winner == "趨勢線":   tl_w   += 1
        elif winner == "MACD":     macd_w += 1
        elif winner == "DCA":      dca_w  += 1

        if winner_i == "平手":     ties_i   += 1
        elif winner_i == "趨勢線": tl_wi    += 1
        elif winner_i == "MACD":   macd_wi  += 1
        elif winner_i == "DCA":    dca_wi   += 1

        rows.append((stock,
                     tl_c, macd_c, dca_c, winner,
                     tl_i, macd_i, dca_i, winner_i))

    for ri, (stock,
             tl_c, macd_c, dca_c, winner,
             tl_i, macd_i, dca_i, winner_i) in enumerate(rows, 2):
        rf = ALT_FILL if ri % 2 == 0 else None

        _cell(ws, ri, 1, stock, fill=rf)

        # ── CAGR 欄（2-4）──
        for col, val, win_key in [(2, tl_c, "趨勢線"), (3, macd_c, "MACD"), (4, dca_c, "DCA")]:
            s      = f"{val:.2f}%" if val is not None else "N/A"
            others = [k for k in ("趨勢線", "MACD", "DCA") if k != win_key]
            f      = RED_FILL  if winner == win_key else (GREY_FILL if winner in others else (TIE_FILL if winner == "平手" else rf))
            fnt    = WIN_FONT  if winner == win_key else (LOSE_FONT if winner in others else NORMAL_FONT)
            _cell(ws, ri, col, s, fill=f, font=fnt)

        # ── CAGR 勝出欄（5）──
        wf  = RED_FILL if winner in ("趨勢線", "MACD", "DCA") else (TIE_FILL if winner == "平手" else rf)
        wfn = WIN_FONT if winner in ("趨勢線", "MACD", "DCA") else NORMAL_FONT
        _cell(ws, ri, 5, winner, fill=wf, font=wfn)

        # ── IRR 欄（6-8）──
        for col, val, win_key in [(6, tl_i, "趨勢線"), (7, macd_i, "MACD"), (8, dca_i, "DCA")]:
            s      = f"{val:.2f}%" if val is not None else "N/A"
            others = [k for k in ("趨勢線", "MACD", "DCA") if k != win_key]
            f      = RED_FILL  if winner_i == win_key else (GREY_FILL if winner_i in others else (TIE_FILL if winner_i == "平手" else rf))
            fnt    = WIN_FONT  if winner_i == win_key else (LOSE_FONT if winner_i in others else NORMAL_FONT)
            _cell(ws, ri, col, s, fill=f, font=fnt)

        # ── IRR 勝出欄（9）──
        wfi  = RED_FILL if winner_i in ("趨勢線", "MACD", "DCA") else (TIE_FILL if winner_i == "平手" else rf)
        wfni = WIN_FONT if winner_i in ("趨勢線", "MACD", "DCA") else NORMAL_FONT
        _cell(ws, ri, 9, winner_i, fill=wfi, font=wfni)

        ws.row_dimensions[ri].height = 20

    _summary_block(ws,
                   start_row=len(rows) + 3,
                   merge_cols="A{r}:I{r}",
                   totals=[("趨勢線勝(CAGR)", tl_w),  ("MACD 勝(CAGR)",  macd_w),
                           ("DCA 勝(CAGR)",   dca_w),  ("平手(CAGR)",     ties)])

    return tl_w, macd_w, dca_w, ties


# ── Sheet 建構：一對一交叉比對 ────────────────────────────────────────────────

def build_duel_sheet(wb, sheet_title, all_stocks,
                     data_a, name_a, data_b, name_b):
    """
    建立單一一對一比對 Sheet。
    data_a / data_b：{stock: cagr_or_irr} 字典
    name_a / name_b：顯示名稱，如 "趨勢線"、"MACD"
    """
    ws   = wb.create_sheet(sheet_title)
    hdr  = [f"{name_a} CAGR", f"{name_b} CAGR", "差距", "勝出"]
    _sheet_header(ws,
                  col_widths=[14, 20, 20, 14, 14],
                  headers=["股票代碼"] + hdr)

    a_wins = b_wins = ties = 0
    rows   = []

    stocks = sorted(set(data_a.keys()) | set(data_b.keys()))
    for stock in stocks:
        a_c = data_a.get(stock)
        b_c = data_b.get(stock)

        if a_c is not None and b_c is not None:
            diff = round(a_c - b_c, 2)
            if a_c > b_c:   winner = name_a; a_wins += 1
            elif b_c > a_c: winner = name_b; b_wins += 1
            else:            winner = "平手"; ties   += 1
        else:
            diff   = None
            winner = "N/A"
        rows.append((stock, a_c, b_c, diff, winner))

    for ri, (stock, a_c, b_c, diff, winner) in enumerate(rows, 2):
        rf = ALT_FILL if ri % 2 == 0 else None

        _cell(ws, ri, 1, stock, fill=rf)

        # 欄 A 策略
        a_str  = f"{a_c:.2f}%" if a_c is not None else "N/A"
        a_fill = RED_FILL  if winner == name_a else (GREY_FILL if winner == name_b else (TIE_FILL if winner == "平手" else rf))
        a_font = WIN_FONT  if winner == name_a else (LOSE_FONT if winner == name_b else NORMAL_FONT)
        _cell(ws, ri, 2, a_str, fill=a_fill, font=a_font)

        # 欄 B 策略
        b_str  = f"{b_c:.2f}%" if b_c is not None else "N/A"
        b_fill = RED_FILL  if winner == name_b else (GREY_FILL if winner == name_a else (TIE_FILL if winner == "平手" else rf))
        b_font = WIN_FONT  if winner == name_b else (LOSE_FONT if winner == name_a else NORMAL_FONT)
        _cell(ws, ri, 3, b_str, fill=b_fill, font=b_font)

        # 差距欄（正 = A 領先，負 = B 領先）
        if diff is not None:
            diff_str  = f"+{diff:.2f}%" if diff > 0 else f"{diff:.2f}%"
            diff_color = "1F7A3A" if diff > 0 else ("CC0000" if diff < 0 else "000000")
            diff_font  = Font(name="Arial", size=11, bold=True, color=diff_color)
            _cell(ws, ri, 4, diff_str, fill=rf, font=diff_font)
        else:
            _cell(ws, ri, 4, "N/A", fill=rf)

        # 勝出欄
        wf  = RED_FILL if winner in (name_a, name_b) else (TIE_FILL if winner == "平手" else rf)
        wfn = WIN_FONT if winner in (name_a, name_b) else NORMAL_FONT
        _cell(ws, ri, 5, winner, fill=wf, font=wfn)
        ws.row_dimensions[ri].height = 20

    _summary_block(ws,
                   start_row=len(rows) + 3,
                   merge_cols="A{r}:E{r}",
                   totals=[(f"{name_a} 勝", a_wins), (f"{name_b} 勝", b_wins), ("平手", ties)])

    total = a_wins + b_wins + ties
    print(f"  [{sheet_title}]  {name_a} 勝 {a_wins}  /  {name_b} 勝 {b_wins}  /  平手 {ties}  (共 {total} 支)")
    return a_wins, b_wins, ties


# ── 主程式 ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  策略對比：趨勢線 vs MACD vs DCA")
    print("=" * 60)

    print(f"\n📂 讀取趨勢線結果：{TL_DIR}")
    tl_data, tl_irr     = scan_folder(TL_DIR, "_交易明細")

    print(f"\n📂 讀取 MACD 結果：{MACD_DIR}")
    macd_data, macd_irr = scan_folder(MACD_DIR, "_MACD_交易明細")

    print(f"\n📂 讀取 DCA 結果：{DCA_DIR}")
    dca_data, dca_irr   = scan_folder(DCA_DIR, "_DCA明細")

    all_stocks = sorted(set(tl_data.keys()) | set(macd_data.keys()) | set(dca_data.keys()))
    if not all_stocks:
        print("\n⚠️  沒有找到任何股票資料，請確認輸出資料夾路徑正確。")
        return

    wb = Workbook()

    # Sheet1：三策略總覽（CAGR + IRR）
    print("\n📊 建立三策略總覽 Sheet ...")
    tl_w, macd_w, dca_w, ties3 = build_overview_sheet(
        wb, all_stocks,
        tl_data, macd_data, dca_data,
        tl_irr,  macd_irr,  dca_irr)

    # Sheet2-4：CAGR 交叉比對
    print("\n📊 建立 CAGR 交叉比對 Sheets ...")
    build_duel_sheet(wb, "趨勢線 vs MACD（CAGR）", all_stocks, tl_data,   "趨勢線", macd_data, "MACD")
    build_duel_sheet(wb, "趨勢線 vs DCA（CAGR）",  all_stocks, tl_data,   "趨勢線", dca_data,  "DCA")
    build_duel_sheet(wb, "MACD vs DCA（CAGR）",    all_stocks, macd_data,  "MACD",   dca_data,  "DCA")

    # Sheet5-7：IRR 交叉比對（純已實現平倉）
    print("\n📊 建立 IRR 交叉比對 Sheets ...")
    build_duel_sheet(wb, "趨勢線 vs MACD（IRR）",  all_stocks, tl_irr,    "趨勢線", macd_irr,  "MACD")
    build_duel_sheet(wb, "趨勢線 vs DCA（IRR）",   all_stocks, tl_irr,    "趨勢線", dca_irr,   "DCA")
    build_duel_sheet(wb, "MACD vs DCA（IRR）",     all_stocks, macd_irr,  "MACD",   dca_irr,   "DCA")

    wb.save(OUTPUT_PATH)

    total3 = tl_w + macd_w + dca_w + ties3
    print(f"\n{'=' * 60}")
    print(f"✅  對比表格已輸出：{OUTPUT_PATH}")
    print(f"\n  📊 三策略總覽（共 {total3} 支股票）")
    print(f"     趨勢線勝：{tl_w} 場")
    print(f"     MACD 勝 ：{macd_w} 場")
    print(f"     DCA 勝  ：{dca_w} 場")
    print(f"     平手    ：{ties3} 場")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
