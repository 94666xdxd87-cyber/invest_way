"""
report.py — Excel 三分頁輸出（每日明細 / 交易記錄 / 績效摘要）。
"""

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    INITIAL_CASH,
    MACD_FAST, MACD_SLOW, MACD_SIGNAL,
    BUY_RATIO, SELL_RATIO,
)


# ── 樣式工具 ──────────────────────────────────────────────────────────────────

def _bdr() -> Border:
    s = Side(style="thin", color="B0C4DE")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row: int, col: int, val, width: float | None = None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", start_color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _bdr()
    if width is not None:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _cell(ws, row: int, col: int, val,
          fill: str | None = None,
          bold: bool = False,
          color: str = "000000",
          align: str = "center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=10, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _bdr()
    if fill:
        c.fill = PatternFill("solid", start_color=fill)
    return c


# ── 主函式 ────────────────────────────────────────────────────────────────────

def save_excel(
    output_path:  str,
    display_name: str,
    daily_df,
    trades:       list,
    perf:         dict,
    prices,
    dates,
) -> None:
    wb            = Workbook()
    closed_trades = [t for t in trades if t["closed"]]

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet1：每日明細
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "每日明細"

    hdrs1   = ["Date", "Close", "MA45", "State", "Signal",
               "目前持有資金", "目前持有股票", "目前損益", "資產總價值"]
    widths1 = [14, 12, 12, 12, 30, 16, 14, 14, 16]
    for ci, (h, w) in enumerate(zip(hdrs1, widths1), 1):
        _hdr(ws1, 1, ci, h, w)
    ws1.row_dimensions[1].height = 20
    ws1.freeze_panes = "A2"

    for ri, row in daily_df.iterrows():
        er      = ri + 2
        sig     = row["Signal"]
        is_buy  = "買入" in sig
        is_sell = "賣出" in sig
        rfill   = ("FFD700"  if is_buy
                   else ("FFCDD2" if is_sell
                         else ("EBF3FB" if ri % 2 == 0 else None)))
        bold    = is_buy or is_sell

        pnl    = row["目前損益"]
        pc     = "1F7A3A" if pnl > 0 else ("CC0000" if pnl < 0 else "000000")
        ta_val = row["資產總價值"]
        ta_c   = "1F7A3A" if ta_val > INITIAL_CASH else ("CC0000" if ta_val < INITIAL_CASH else "000000")

        vals1 = [row["Date"], row["Close"], row["MA45"], row["State"], sig]
        alns1 = ["center", "right", "right", "center", "left"]
        for ci, (v, a) in enumerate(zip(vals1, alns1), 1):
            _cell(ws1, er, ci, v, fill=rfill, bold=bold, align=a)
        _cell(ws1, er, 6, row["目前持有資金"], fill=rfill, bold=bold, align="right")
        _cell(ws1, er, 7, row["目前持有股票"], fill=rfill, bold=bold, align="right")
        _cell(ws1, er, 8, pnl,                fill=rfill, bold=bold, color=pc,   align="right")
        _cell(ws1, er, 9, ta_val,             fill=rfill, bold=bold, color=ta_c, align="right")

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet2：交易記錄
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("交易記錄")

    hdrs2   = ["#", "買入日期", "買入價", "賣出日期", "賣出價", "股數",
               "投入成本", "賣出收入", "損益", "報酬率%", "持有天數", "狀態",
               "持有資金", "持有股票", "剩餘持倉浮動損益"]
    widths2 = [5, 14, 12, 14, 12, 8, 14, 14, 14, 12, 12, 10, 14, 12, 18]
    for ci, (h, w) in enumerate(zip(hdrs2, widths2), 1):
        _hdr(ws2, 1, ci, h, w)
    ws2.row_dimensions[1].height = 20
    ws2.freeze_panes = "A2"

    for ri, t in enumerate(trades, 1):
        er   = ri + 1
        pnl  = t["pnl"]
        fill = ("C8E6C9" if pnl > 0 else ("FFCDD2" if pnl < 0 else None))
        uc   = ("1F7A3A" if t["unrealized_pnl"] > 0
                else ("CC0000" if t["unrealized_pnl"] < 0 else "000000"))

        vals2 = [
            t["trade_no"],
            t["buy_date"].strftime("%Y-%m-%d"),
            t["buy_price"],
            t["sell_date"].strftime("%Y-%m-%d"),
            t["sell_price"],
            t["shares"],
            t["cost"],
            t["revenue"],
            t["pnl"],
            f"{t['ret_pct']:.2f}%",
            t["hold_days"],
            "已平倉" if t["closed"] else "★持倉中",
        ]
        alns2 = ["center","center","right","center","right","center",
                 "right","right","right","right","center","center"]
        for ci, (v, a) in enumerate(zip(vals2, alns2), 1):
            _cell(ws2, er, ci, v, fill=fill, align=a)
        _cell(ws2, er, 13, t["cash_after"],     fill=fill, align="right")
        _cell(ws2, er, 14, t["shares_after"],   fill=fill, align="right")
        _cell(ws2, er, 15, t["unrealized_pnl"], fill=fill, color=uc, align="right")

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet3：績效摘要
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("績效摘要")
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 22

    irr_val    = perf["irr_val"]
    irr_closed = perf["irr_closed_only"]
    cagr       = perf["cagr"]
    mdd_str    = f"{perf['max_drawdown'] * 100:.2f}%"
    cr_str     = f"{perf['cumulative_return'] * 100:.2f}%"
    cagr_str   = f"{cagr * 100:.2f}%" if cagr is not None else "N/A"

    summary = [
        ("初始資金",                        INITIAL_CASH),
        ("策略期間",                        perf["date_range"]),
        ("", ""),
        ("─── 傳統 MACD 策略參數 ───",     ""),
        ("MACD 週期（快/慢/信號）",         f"{MACD_FAST} / {MACD_SLOW} / {MACD_SIGNAL}"),
        ("買入規則",                        "金叉（MACD 線上穿 Signal 線）"),
        ("賣出規則",                        "死叉（MACD 線下穿 Signal 線）"),
        ("買入動用現金比例",               f"{BUY_RATIO * 100:.0f}%"),
        ("賣出持股比例",                   f"{SELL_RATIO * 100:.0f}%"),
        ("", ""),
        ("─── 交易統計 ───",              ""),
        ("總交易筆數（含未平倉）",          len(trades)),
        ("已平倉筆數",                      len(closed_trades)),
        ("未平倉筆數（期末持倉）",          len(trades) - len(closed_trades)),
        ("已平倉獲利筆數",                  perf["win_trades"]),
        ("已平倉虧損筆數",                  sum(1 for t in closed_trades if t["pnl"] <= 0)),
        ("已平倉勝率",                      f"{perf['win_rate']:.1f}%"),
        ("", ""),
        ("─── 已實現損益（僅平倉）───",   ""),
        ("已實現總投入",                    round(sum(t["cost"]    for t in closed_trades), 2)),
        ("已實現總收入",                    round(sum(t["revenue"] for t in closed_trades), 2)),
        ("已實現總損益",                    perf["realized_pnl"]),
        ("", ""),
        ("─── 含未實現損益（期末帳戶）───", ""),
        ("期末持有資金",                    perf["final_cash"]),
        ("期末持有股票",                    perf["final_shares"]),
        ("期末帳戶總損益",                  perf["final_pnl"]),
        ("期末資產總價值",                  perf["final_total_asset"]),
        ("累積報酬率（總，非年化）",        cr_str),
        ("最大回撤 MDD",                    mdd_str),
        ("", ""),
        ("─── 年化報酬率 ───",            ""),
        ("CAGR（複合年化成長率）",          cagr_str),
        ("年化 IRR（含未平倉假設清算）",    f"{irr_val}%"    if irr_val    is not None else "N/A"),
        ("年化 IRR（純已實現平倉）",        f"{irr_closed}%" if irr_closed is not None else "N/A"),
        ("", ""),
        ("─── 風險調整報酬 ───",           ""),
        ("Sortino Ratio（無風險利率 1.7%）",
            perf["sortino_ratio"] if perf.get("sortino_ratio") is not None else "N/A"),
        ("", ""),
        ("─── 已平倉單筆統計 ───",        ""),
        ("平均持有天數",
            round(np.mean([t["hold_days"] for t in closed_trades]), 1) if closed_trades else 0),
        ("最大單筆獲利",
            max((t["pnl"] for t in closed_trades), default=0)),
        ("最大單筆虧損",
            min((t["pnl"] for t in closed_trades), default=0)),
        ("平均單筆損益",
            round(np.mean([t["pnl"] for t in closed_trades]), 2) if closed_trades else 0),
    ]

    title_c = ws3.cell(row=1, column=1, value=f"{display_name}  傳統 MACD 策略  績效摘要")
    title_c.font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
    title_c.alignment = Alignment(horizontal="center")
    ws3.merge_cells("A1:B1")
    ws3.row_dimensions[1].height = 28

    section_fill = "D6E4F0"
    for ri, (label, val) in enumerate(summary, 2):
        if not label:
            continue
        is_section = label.startswith("───")

        lc = ws3.cell(row=ri, column=1, value=label)
        lc.font      = Font(name="Arial", bold=True, size=10,
                            color="1F4E79" if is_section else "000000")
        lc.fill      = PatternFill("solid",
                                   start_color=section_fill if is_section else "EBF3FB")
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border    = _bdr()

        if is_section:
            vc = ws3.cell(row=ri, column=2, value="")
            vc.fill   = PatternFill("solid", start_color=section_fill)
            vc.border = _bdr()
            ws3.row_dimensions[ri].height = 16
            continue

        if isinstance(val, str) and val not in ("N/A", ""):
            try:
                num = float(val.replace("%", ""))
                tc  = "1F7A3A" if num > 0 else ("CC0000" if num < 0 else "000000")
            except ValueError:
                tc = "000000"
        elif isinstance(val, (int, float)):
            tc = "1F7A3A" if val > 0 else ("CC0000" if val < 0 else "000000")
        else:
            tc = "000000"

        vc = ws3.cell(row=ri, column=2, value=val)
        vc.font      = Font(name="Arial", size=10, color=tc)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border    = _bdr()
        ws3.row_dimensions[ri].height = 18

    wb.save(output_path)
    print(f"  ✅ Excel → {output_path}")
